"""
학생 엑셀 Import/Export 서비스

배포환경 고려:
- openpyxl이 설치되지 않은 환경에서는 기능을 명확한 오류로 안내(앱 크래시 방지)
- 컬럼 검증/전화번호 정규화/상태 값 정규화 지원
"""

from __future__ import annotations

import re
from typing import Dict, List, Tuple

from core.models import Student


REQUIRED_STUDENT_COLUMNS = [
    "학년",
    "상태",
    "학생이름",
    "학교명",
    "학부모 연락처",
    "학생 연락처",
]


def normalize_phone(v: str) -> str:
    s = (v or "").strip()
    if not s:
        return ""
    digits = re.sub(r"[^0-9]", "", s)
    if len(digits) == 11 and digits.startswith("01"):
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 10 and digits.startswith("01"):
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    # 폴백: 원본 유지
    return s


def _normalize_status(v: str) -> str:
    s = (v or "").strip()
    if s in ("재원", "휴원", "퇴원"):
        return s
    # 흔한 변형 처리
    if s in ("재학", "재원중", "재원생"):
        return "재원"
    if s in ("휴학", "휴원중"):
        return "휴원"
    if s in ("전원", "퇴학", "퇴원중"):
        return "퇴원"
    return s or "재원"


def _openpyxl():
    try:
        import openpyxl  # type: ignore

        return openpyxl
    except Exception as e:
        raise RuntimeError(
            "엑셀 기능을 사용하려면 openpyxl 설치가 필요합니다.\n\n"
            "설치: pip install openpyxl\n\n"
            f"원인: {type(e).__name__}: {e}"
        )


def export_students_to_xlsx(students: List[Student], output_path: str) -> None:
    openpyxl = _openpyxl()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "학생"

    ws.append(list(REQUIRED_STUDENT_COLUMNS))
    for s in students:
        ws.append(
            [
                (s.grade or "").strip(),
                (s.status or "").strip(),
                (s.name or "").strip(),
                (s.school_name or "").strip(),
                (s.parent_phone or "").strip(),
                (s.student_phone or "").strip(),
            ]
        )

    # 간단 가독성: 컬럼 폭
    widths = {
        "A": 8,   # 학년
        "B": 8,   # 상태
        "C": 14,  # 학생이름
        "D": 18,  # 학교명
        "E": 16,  # 학부모 연락처
        "F": 16,  # 학생 연락처
    }
    try:
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
    except Exception:
        pass

    wb.save(output_path)


def import_students_from_xlsx(input_path: str) -> Tuple[List[Student], Dict[str, int]]:
    """
    Returns:
        (students, stats)
        stats = {"rows": n, "imported": n, "skipped": n}
    """
    openpyxl = _openpyxl()
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # header
    header = []
    for cell in ws[1]:
        header.append(str(cell.value or "").strip())

    # 헤더 alias 허용
    alias = {
        "학생 이름": "학생이름",
        "학생명": "학생이름",
        "학부모연락처": "학부모 연락처",
        "학생연락처": "학생 연락처",
        "학교": "학교명",
    }
    header = [alias.get(h, h) for h in header]

    idx = {h: i for i, h in enumerate(header)}
    missing = [c for c in REQUIRED_STUDENT_COLUMNS if c not in idx]
    if missing:
        raise ValueError(f"엑셀 컬럼이 부족합니다: {', '.join(missing)}\n필수 컬럼: {', '.join(REQUIRED_STUDENT_COLUMNS)}")

    out: List[Student] = []
    rows = 0
    skipped = 0

    for r in ws.iter_rows(min_row=2, values_only=True):
        rows += 1
        def _get(col: str) -> str:
            v = r[idx[col]] if idx[col] < len(r) else ""
            return str(v or "").strip()

        grade = _get("학년")
        status = _normalize_status(_get("상태"))
        name = _get("학생이름")
        school_name = _get("학교명")
        parent_phone = normalize_phone(_get("학부모 연락처"))
        student_phone = normalize_phone(_get("학생 연락처"))

        # 완전 빈 줄 스킵
        if not any([grade, status, name, school_name, parent_phone, student_phone]):
            skipped += 1
            continue
        if not name:
            skipped += 1
            continue

        out.append(
            Student(
                grade=grade,
                status=status,
                name=name,
                school_name=school_name,
                parent_phone=parent_phone,
                student_phone=student_phone,
            )
        )

    return out, {"rows": rows, "imported": len(out), "skipped": skipped}

